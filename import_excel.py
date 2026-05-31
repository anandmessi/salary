import openpyxl
from database import upsert_worker
from schema import Worker


def import_employees(filepath='employee.xlsx') -> dict:
    """
    Import workers from Excel file.
    Returns {"imported": int, "skipped": int, "errors": list[str]}
    """
    result = {"imported": 0, "skipped": 0, "errors": []}

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except FileNotFoundError:
        result["errors"].append(f"File not found: {filepath}")
        return result
    except Exception as e:
        result["errors"].append(f"Cannot open file: {e}")
        return result

    ws = wb.active
    if ws.max_row < 3:
        result["errors"].append("Excel file has no data rows (expected data from row 3)")
        return result

    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        try:
            sl_no = row[0] if len(row) > 0 else None
            name  = row[1] if len(row) > 1 else None
            ifsc  = row[2] if len(row) > 2 else None
            acct  = row[3] if len(row) > 3 else None

            if not sl_no and not name:
                result["skipped"] += 1
                continue

            if not name or not str(name).strip():
                result["errors"].append(f"Row {row_num}: Missing worker name")
                result["skipped"] += 1
                continue

            if sl_no is None:
                result["errors"].append(f"Row {row_num}: Missing serial number for '{name}'")
                result["skipped"] += 1
                continue

            worker_id = (
                f"EMP{int(sl_no):03d}"
                if isinstance(sl_no, (int, float))
                else str(sl_no).strip()
            )

            w = Worker(
                worker_id=worker_id,
                name=str(name).strip(),
                ifsc_code=str(ifsc).strip() if ifsc else '',
                bank_account=str(acct).strip() if acct else '',
                skill_category='Unskilled',
                active=True,
            )
            upsert_worker(w)
            result["imported"] += 1

        except Exception as e:
            result["errors"].append(f"Row {row_num}: {e}")
            result["skipped"] += 1

    return result


if __name__ == '__main__':
    res = import_employees()
    print(f"Imported: {res['imported']}, Skipped: {res['skipped']}")
    if res['errors']:
        print("Errors:")
        for err in res['errors']:
            print(f"  {err}")
